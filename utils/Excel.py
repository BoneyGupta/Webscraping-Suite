import time

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import xlwings as xw


class Excel:
    def __init__(self, file_path: str, data_only: bool):
        self.file_path = file_path
        if data_only:
            self.workbook = load_workbook(filename=self.file_path, data_only=True)
        else:
            self.workbook = load_workbook(filename=self.file_path, data_only=False)
        self.current_sheet = self.workbook.active

    def get_sheet_count(self) -> int:
        return len(self.workbook.sheetnames)

    def get_sheet_names(self):
        sheet_names = self.workbook.sheetnames
        return sheet_names

    def get_current_sheet_name(self) -> str:
        return self.current_sheet.title

    def change_sheet(self, sheet_index: int) -> None:
        if 0 <= sheet_index < self.get_sheet_count():
            self.current_sheet = self.workbook.worksheets[sheet_index]
        else:
            print(f"Invalid sheet index: {sheet_index}. Available range: 0 to {self.get_sheet_count() - 1}")

    def get_rows_count(self) -> int:
        return self.current_sheet.max_row

    def get_columns_count(self) -> int:
        return self.current_sheet.max_column

    def get_cell_value(self, row: int, column: int):
        return self.current_sheet.cell(row=row, column=column).value

    def is_cell_empty(self, row: int, column: int) -> bool:
        return self.get_cell_value(row, column) is None

    def get_cell_data_type(self, row: int, column: int):
        value = self.get_cell_value(row, column)
        if value is not None:
            return type(value)
        return None

    def enter_string(self, row: int, column: int, value: str) -> None:
        cell = self.current_sheet.cell(row=row, column=column)
        cell.value = value

    def enter_string_with_green_font(self, row: int, column: int, value: str) -> None:
        cell = self.current_sheet.cell(row=row, column=column)
        cell.value = value
        cell.font = Font(color="00FF00")  # Green color

    def enter_string_with_red_font(self, row: int, column: int, value: str) -> None:
        cell = self.current_sheet.cell(row=row, column=column)
        cell.value = value
        cell.font = Font(color="FF0000")  # Red color

    def enter_string_with_green_cell_highlight(self, row: int, column: int, value: str) -> None:
        cell = self.current_sheet.cell(row=row, column=column)
        cell.value = value
        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green fill

    def enter_string_with_red_cell_highlight(self, row: int, column: int, value: str) -> None:
        cell = self.current_sheet.cell(row=row, column=column)
        cell.value = value
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red fill

    def save_workbook(self) -> None:
        self.workbook.save(filename=self.file_path)

    def close_workbook(self) -> None:
        self.workbook.close()

    def recalculate_xl(self):
        # Open the workbook in Excel
        wb = xw.Book(self.file_path)

        # Recalculate the entire workbook
        wb.app.calculate()

        # Save the workbook with recalculated values
        wb.save(self.file_path)
        time.sleep(2)

        wb.app.quit()
        time.sleep(2)
